import sys
import json
import time
import random
import os.path
import threading
import pystray
import datetime
from PIL import Image
from pathlib import Path
from openpyxl import load_workbook
from PyQt5.QtGui import QMovie, QCursor, QFont, QGuiApplication
from PyQt5.QtCore import Qt, QTimer,QUrl
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QSizePolicy, QLineEdit, QPushButton, QGridLayout,QMenu,QVBoxLayout,QAction,QDialog

VOICE_DIR_ERROR='voice_dir_error'
DEFAULT_VOICE_LENGTH=2500
CLICK_SHREDHOLD=110

def get_action(gif_dir, action, action_dir):
    return [os.path.join(gif_dir,action,it) for it in action_dir]

def get_word(voice_dir,worksheet,idx1,idx2,duration=DEFAULT_VOICE_LENGTH):
    ret_word=[]
    error='error'
    for row in worksheet.iter_rows(min_row=2,values_only=True):
        if row and row[idx1] is not None:
            if row[idx2] is None:
                ret_word.append((row[idx1].strip(),VOICE_DIR_ERROR,duration))
            else:
                ret_word.append((row[idx1].strip(),os.path.join(voice_dir,row[idx2]),duration))
    return ret_word

def read_config(config_file):
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
                return config
        except json.JSONDecodeError:
            pass
    return {}

def save_config(config,config_file):
    with open(config_file, 'w') as f:
        json.dump(config, f)

class PathData:
    def __init__(self):
        self.current_dir = Path(__file__).resolve().parent
        self.gif_dir = self.current_dir / 'source' / '动作gif格式'
        self.voice_dir = self.current_dir / 'source' / '播放语音素材'
        self.config_file = self.current_dir / 'source' / '数据记录' /'config.json'
        if not self.gif_dir.exists() or not self.voice_dir.exists():
            raise FileNotFoundError("指定的素材文件夹不存在，请检查路径。")
        self.born_gif_dir = self.get_dir('出生动作')
        self.drag_gif_dir = self.get_dir('拖拽动作')
        self.idle_gif_dir = self.get_dir('待机动作')
        self.walk_gif_dir = self.get_dir('行走动作')
        self.fall_gif_dir = self.get_dir('降落动作')
        self.click_gif_dir = self.get_dir('点击动作')
        voice_excel = self.voice_dir / '语音目录.xlsx'
        if not voice_excel.exists():
            raise FileNotFoundError("语音目录.xlsx 文件不存在，请检查路径。")
        workbook = load_workbook(voice_excel)
        self.worksheet = workbook['Sheet1']
    def get_dir(self, action):
        action_path = self.gif_dir / action
        if action_path.exists():
            return [file.name for file in action_path.iterdir() if file.is_file()]
        return []

class GifData:
    def __init__(self,pathdata):
        self.drag_gif = get_action(pathdata.gif_dir,'拖拽动作',pathdata.drag_gif_dir)
        self.fall_gif = get_action(pathdata.gif_dir,'降落动作',pathdata.fall_gif_dir)
        self.walk_gif = get_action(pathdata.gif_dir,'行走动作',pathdata.walk_gif_dir)
        self.born_gif = get_action(pathdata.gif_dir,'出生动作',pathdata.born_gif_dir)
        self.idle_gif = get_action(pathdata.gif_dir,'待机动作',pathdata.idle_gif_dir)
        self.click_gif= get_action(pathdata.gif_dir,'点击动作',pathdata.click_gif_dir)

class WordData:
    def __init__(self,voice_dir,worksheet):
        self.idle_word             = get_word(voice_dir,worksheet,0,1)
        self.pre_morning_idle_word = get_word(voice_dir,worksheet,2,3)
        self.pre_noon_idle_word    = get_word(voice_dir,worksheet,4,5)
        self.pre_night_idle_word   = get_word(voice_dir,worksheet,6,7)
        self.click_word       = get_word(voice_dir,worksheet,8,9)
        self.drag_word        = get_word(voice_dir,worksheet,10,11)
        self.fall_word        = get_word(voice_dir,worksheet,12,13)
        self.born_word        = get_word(voice_dir,worksheet,14,15)
    def get_morning_idle_word(self):
        return self.idle_word + self.pre_morning_idle_word
    def get_noon_idle_word(self):
        return self.idle_word + self.pre_noon_idle_word
    def get_night_idle_word(self):
        return self.idle_word + self.pre_night_idle_word

class Status:
    def __init__(self):
        self.is_dragging = False
        self.is_falling = False
        self.is_walking = False
        self.is_muted=False
        self.is_clicking=False
        self.is_inputing=False
        self.is_random_voice=True

class Parameters:
    def __init__(self):
        self.walk_direction = -1
        self.walk_speed = 2
        self.fall_speed = 5
        self.fall_height=250
        self.land_height=5
        self.volume=50
        self.reserve_space=0
        self.reserve_space2=0
        self.timer2_interval=5000
        self.timer1_interval=30
        self.morning_start = 5
        self.morning_end = 12
        self.noon_start = 12
        self.noon_end = 18

class SpecialGif:
    def __init__(self):
        self.drag_gif=''
        self.fall_gif=''
        self.walk_gif=''

class SpecialWord:
    def __init__(self):
        self.drag_word=''
        self.fall_word=''

class MemoryDay:
    def __init__(self):
        self.last_date=[0,0,0]
        self.day_count=0
        self.today=[0,0,0]

    def get_accompany_days(self):
        return self.day_count

class DesktopPet(QWidget):
    #初始化
    def __init__(self):
        super().__init__()
        self.path_data=PathData() #数据
        self.gif_data=GifData(self.path_data) #gif数据
        self.word_data=WordData(self.path_data.voice_dir,self.path_data.worksheet) #音频数据
        self.status=Status() #状态
        self.parameters=Parameters() #参数
        self.special_gif=SpecialGif() #特殊gif
        self.special_word=SpecialWord() #特殊音频
        self.config_file=self.path_data.config_file
        self.current_dir=self.path_data.current_dir
        del self.path_data

        self.start_time=time.time()
        self.elapsed_time=0
        self.prev_state = None
        self.prev_mouse_pos = None
        self.mouse_is_pressed = False
        self.dialog_label=None

        self.memory_day=MemoryDay()

        self.json_setting()

        self.update_day_count()

        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground)

        #音频播放器
        self.player = QMediaPlayer(self)
        self.player.setVolume(self.parameters.volume)
        self.player.mediaStatusChanged.connect(self.handle_media_status)

        #动画刷新Timer1
        self.timer1 = QTimer(self)
        self.timer1.timeout.connect(self.update_animation)
        self.timer1.start(self.parameters.timer1_interval)

        #随机语音Timer2
        self.timer2 = QTimer(self)
        self.restart_timer2()
        self.timer2.timeout.connect(self.random_conversation)
        self.record_start_time()

        self.timer3 = QTimer(self)
        self.timer3.setSingleShot(True)
        self.timer3.timeout.connect(self.differ_drag_and_click)

        #出生动画设置
        self.move(random.randint(0, QApplication.desktop().screenGeometry().width() - self.width()),
                         random.randint(0, QApplication.desktop().screenGeometry().height() - self.height()))
        self.current_movie = self.load_gif(random.choice(self.gif_data.born_gif))  #出生动画
        self.current_movie.frameChanged.connect(self.check_last_frame)
        self.label = QLabel(self)
        self.label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.label.setMovie(self.current_movie)
        self.label.setGeometry(0, 0, self.current_movie.currentPixmap().width() + self.parameters.reserve_space2,
                               self.current_movie.currentPixmap().height())
        self.current_movie.start()
        self.update()

        #出生随机语音设置
        self.full_speak(random.choice(self.word_data.born_word))

    def update_day_count(self):
        now=datetime.datetime.now()
        self.memory_day.today=[now.year,now.month,now.day]
        if self.memory_day.today != self.memory_day.last_date:
            self.memory_day.day_count+=1
            self.memory_day.last_date=self.memory_day.today
            config=read_config(self.config_file)
            config['LAST_DATE']=self.memory_day.last_date
            config['DAY_COUNT']=self.memory_day.day_count
            save_config(config,self.config_file)

    def getday(self):
        now = datetime.datetime.now()
        hour = now.hour
        hour = hour + now.minute / 60
        if self.parameters.morning_start <= hour < self.parameters.morning_end:
            return 0
        elif self.parameters.noon_start <= hour < self.parameters.noon_end:
            return 1
        elif hour < self.parameters.morning_start or hour > self.parameters.noon_end:
            return 2
        else:
            return 3

    def json_setting(self):
        config=read_config(self.config_file)
        if 'MUTE' in config:
            self.status.is_muted=config['MUTE']
        if 'RANDOM_VOICE' in config:
            self.status.is_random_voice=config['RANDOM_VOICE']
        if 'VOLUME' in config:
            self.parameters.volume=config['VOLUME']
        if 'INTERVAL' in config:
            self.parameters.timer2_interval=config['INTERVAL']
        if 'FALL_HEIGHT' in config:
            self.parameters.fall_height=config['FALL_HEIGHT']
        if 'LAND_HEIGHT' in config:
            self.parameters.land_height=config['LAND_HEIGHT']
        if 'MORNING_START' in config:
            self.parameters.morning_start=config['MORNING_START']
        if 'MORNING_END' in config:
            self.parameters.morning_end=config['MORNING_END']
        if 'NOON_START' in config:
            self.parameters.noon_start=config['NOON_START']
        if 'NOON_END' in config:
            self.parameters.noon_end=config['NOON_END']
        if 'LAST_DATE' in config:
            self.memory_day.last_date=config['LAST_DATE']
        if 'DAY_COUNT' in config:
            self.memory_day.day_count=config['DAY_COUNT']

    def record_start_time(self):
        self.start_time=time.time()

    def stop_timer2(self):
        current_time=time.time()
        self.elapsed_time=int((current_time-self.start_time)*1000)%self.parameters.timer2_interval
        self.timer2.stop()

    def restart_timer2(self):
        if self.status.is_random_voice:
            self.timer2.timeout.connect(self.restore_interval)
            self.timer2.start(self.parameters.timer2_interval-self.elapsed_time)

    def restore_interval(self):
        self.timer2.setInterval(self.parameters.timer2_interval)
        self.timer2.timeout.disconnect(self.restore_interval)

    def handle_media_status(self,status):
        if status==QMediaPlayer.EndOfMedia:
            self.player.stop()
            self.player.setMedia(QMediaContent())

    def display_word(self,word):
        if self.dialog_label is None:
            self.dialog_label = QLabel()
            self.dialog_label.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint| Qt.Tool)
            self.dialog_label.setAttribute(Qt.WA_TranslucentBackground)
            self.dialog_label.setStyleSheet("""
                        background-color: rgba(0, 0, 0, 0);
                        color: white;
                        font-family: "Segoe UI", sans-serif;
                        font-size: 12pt;
                        font-weight: bold;
                    """)
            self.dialog_label.setAlignment(Qt.AlignCenter)
            self.dialog_label.setWordWrap(True)
            self.dialog_label.setVisible(False)  # 默认不显示
        self.dialog_label.setText(word)
        self.dialog_label.adjustSize()
        self.label_move()
        self.dialog_label.setVisible(True)

    def play_word(self,word_dir):
        if not self.status.is_muted:
            if word_dir != VOICE_DIR_ERROR:
                if self.status.is_random_voice:
                    self.stop_timer2()
                media=QMediaContent(QUrl.fromLocalFile(word_dir))
                if media is not None:
                    self.player.setMedia(media)
                    self.player.play()
                del media

    def vanish_word(self,duration):
        if duration!=0:
            QTimer.singleShot(duration, self.hide_conversation)
        else:
            self.hide_conversation()
        if not self.status.is_clicking and not self.status.is_dragging and not self.status.is_falling:
            self.elapsed_time=0
        QTimer.singleShot(duration,self.restart_timer2)

    def label_move(self):
        if self.dialog_label is not None:
            x = self.pos().x() + (self.width() - self.dialog_label.width()) // 2
            y = self.pos().y() - self.dialog_label.height()
            self.dialog_label.move(x, y)

    def moveEvent(self,event):
        self.label_move()
        super().moveEvent(event)

    def full_speak(self,word):
        self.display_word(word[0])
        self.play_word(word[1])
        self.vanish_word(word[2])

    def hide_conversation(self):
        if self.dialog_label is not None:
            self.dialog_label.setVisible(False)
            del self.dialog_label
            self.dialog_label=None

    def random_conversation(self):
        self.start_time=time.time()
        if not self.status.is_dragging and not self.status.is_falling:
            day=self.getday()
            if day == 0:
                word_voice = random.choice(self.word_data.get_morning_idle_word())
            elif day == 1:
                word_voice = random.choice(self.word_data.get_noon_idle_word())
            elif day==2:
                word_voice = random.choice(self.word_data.get_night_idle_word())
            else:
                word_voice=random.choice(self.word_data.idle_word)
            self.full_speak(word_voice)

    def load_gif(self, gif_path):
        movie = QMovie(gif_path)
        movie.setCacheMode(QMovie.CacheAll)
        movie.start()
        self.setFixedSize(movie.currentPixmap().width() + self.parameters.reserve_space2,
        movie.currentPixmap().height() + self.parameters.reserve_space)
        return movie

    def stop_movie(self):
        self.prev_state = self.get_current_state()
        self.current_movie.stop()
        del self.current_movie
        self.current_movie=None
        self.label.clear()

    def load_movie(self,new_movie):
        self.current_movie=self.load_gif(new_movie)

    def start_movie(self):
        self.current_movie.frameChanged.connect(self.restore_previous_state)
        self.label.setMovie(self.current_movie)
        self.label.setGeometry(0, 0, self.current_movie.currentPixmap().width() + self.parameters.reserve_space2,
                               self.current_movie.currentPixmap().height() + self.parameters.reserve_space)
        self.current_movie.start()

    def switch_movie(self,action_gif):
        self.stop_movie()
        current_pos = self.pos()
        left_bottom_y = current_pos.y() + self.height()
        self.load_movie(action_gif)
        self.start_movie()
        new_y = left_bottom_y - self.height()
        self.move(current_pos.x(), new_y)
        self.label_move()
        self.update()

    def check_last_frame(self, frame_number):
        if frame_number == self.current_movie.frameCount() - 1 and not self.status.is_falling and not self.status.is_walking:
                self.switch_movie(random.choice(self.gif_data.idle_gif))

    def restore_previous_state(self, frame_number):
        if frame_number == self.current_movie.frameCount() - 1:
            if self.prev_state == 'input':
                new_movie=self.special_gif.input_gif
            elif self.prev_state == 'idle':
                new_movie = random.choice(self.gif_data.idle_gif)
            elif self.prev_state == 'drag':
                new_movie = self.special_gif.drag_gif
            elif self.prev_state == 'fall':
                new_movie = self.special_gif.fall_gif
            elif self.prev_state == 'walk':
                new_movie = self.special_gif.walk_gif
            self.switch_movie(new_movie)

    def get_current_state(self):
        if self.status.is_inputing:
            return 'input'
        elif self.status.is_dragging:
            return 'drag'
        elif self.status.is_falling:
            return 'fall'
        elif self.status.is_walking:
            return 'walk'
        else:
            return 'idle'

    def press_to_drag(self):
        self.status.is_walking = False
        self.status.is_falling = False
        self.status.is_clicking = False
        self.status.is_dragging = True
        self.setCursor(QCursor(Qt.OpenHandCursor))
        self.prev_mouse_pos = QCursor.pos()
        self.special_gif.drag_gif = random.choice(self.gif_data.drag_gif)

        self.stop_movie()
        self.load_movie(self.special_gif.drag_gif)
        self.move(QCursor.pos().x()-self.width()//2,QCursor().pos().y()-self.height()//2)
        self.start_movie()
        self.update()

        self.special_word.drag_word = random.choice(self.word_data.drag_word)
        self.play_word(self.special_word.drag_word[1])
        self.display_word(self.special_word.drag_word[0])

    def press_to_click(self):
        self.status.is_walking = False
        self.status.is_falling = False
        self.status.is_dragging = False
        self.status.is_clicking = True

    def release_cancel_click(self):
        self.full_speak(random.choice(self.word_data.click_word))
        self.status.is_clicking=False
        self.switch_movie(random.choice(self.gif_data.click_gif))

    def release_to_fall(self):
        self.status.is_walking = False
        self.status.is_clicking = False
        self.status.is_dragging = False
        self.status.is_falling = True
        self.special_gif.fall_gif = random.choice(self.gif_data.fall_gif)
        self.special_word.fall_word = random.choice(self.word_data.fall_word)
        self.play_word(self.special_word.fall_word[1])
        self.display_word(self.special_word.fall_word[0])
        self.switch_movie(self.special_gif.fall_gif)

    def release_to_idle(self):
        self.status.is_walking = False
        self.status.is_clicking = False
        self.status.is_dragging = False
        self.status.is_falling = False
        self.switch_movie(random.choice(self.gif_data.idle_gif))

    def release_cancel_drag(self):
        self.setCursor(QCursor(Qt.ArrowCursor))
        self.prev_mouse_pos = None
        self.vanish_word(0)
        self.status.is_dragging = False

    def update_fall_to_walk(self):
        self.status.is_clicking = False
        self.status.is_dragging = False
        self.status.is_walking = True
        self.vanish_word(0)
        self.status.is_falling = False
        self.special_gif.walk_gif = random.choice(self.gif_data.walk_gif)
        self.switch_movie(self.special_gif.walk_gif)

    def update_walk_to_idle(self):
        self.status.is_walking = False
        self.switch_movie(random.choice(self.gif_data.idle_gif))

    def differ_drag_and_click(self):
        if self.mouse_is_pressed:
            self.press_to_drag()
        else:
            self.press_to_click()
            self.release_cancel_click()

    def move_mouse(self,current_mouse_pos):
        new_pos = self.pos() + current_mouse_pos - self.prev_mouse_pos
        screens = QApplication.screens()
        combined_geometry = screens[0].geometry()
        for screen in screens[1:]:
            combined_geometry = combined_geometry.united(screen.geometry())
        new_x = max(combined_geometry.left(), min(new_pos.x(), combined_geometry.right() - self.width()))
        new_y = max(combined_geometry.top(), min(new_pos.y(), combined_geometry.bottom() - self.height()))
        self.move(new_x, new_y)
        self.prev_mouse_pos = current_mouse_pos  # 更新上一次鼠标位置

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.mouse_is_pressed=True
            self.timer3.start(CLICK_SHREDHOLD)
        elif event.button() == Qt.RightButton:
            self.press_to_menu()

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.mouse_is_pressed=False
            if self.status.is_dragging:
                self.release_cancel_drag()
                if self.pos().y() < self.parameters.fall_height:
                    self.release_to_fall()
                elif self.status.is_inputing:
                    self.switch_movie(self.special_gif.input_gif)
                else:
                    self.release_to_idle()

    def update_animation(self):
        if self.status.is_dragging:
            self.move_mouse(QCursor.pos())

        elif self.status.is_falling and not self.status.is_dragging:
            new_y = self.pos().y() + self.parameters.fall_speed
            if new_y >= QApplication.desktop().screenGeometry().height() - self.height() -self.parameters.land_height:
                self.update_fall_to_walk()
            else:
                self.move(self.pos().x(), new_y)
                self.update()

        elif self.status.is_walking:
            new_x = self.pos().x() + self.parameters.walk_speed * self.parameters.walk_direction
            if new_x <= 20 or self.pos().y() < QApplication.desktop().screenGeometry().height() - self.height() - self.parameters.land_height-45:
                self.update_walk_to_idle()
            else:
                self.move(new_x, self.pos().y())
                self.update()

    def set_volume(self,volume):
        if volume != '':
            self.parameters.volume = int(volume)
            self.player.setVolume(self.parameters.volume)

    def set_interval(self,interval):
        if interval != '':
            self.parameters.timer2_interval=int(interval)*1000
            if self.timer2.isActive():
                self.timer2.start(self.parameters.timer2_interval)

    def set_fall_height(self,fall_height):
        if fall_height != '':
            self.parameters.fall_height=int(fall_height)

    def set_land_height(self,land_height):
        if land_height != '':
            self.parameters.land_height=int(land_height)

    def set_morning_start(self,morning_start):
        if morning_start!='':
            self.parameters.morning_start=float(morning_start)

    def set_morning_end(self,morning_end):
        if morning_end!='':
            self.parameters.morning_end=float(morning_end)

    def set_noon_start(self,noon_start):
        if noon_start!='':
            self.parameters.noon_start=float(noon_start)

    def set_noon_end(self,noon_end):
        if noon_end!='':
            self.parameters.noon_end=float(noon_end)

    def parameters_text(self):
        return '调节参数'

    def parameters_setting(self):
        labels=[
            ('音量', str(self.parameters.volume),'语音间隔(s)',str(int(self.parameters.timer2_interval/1000))),
            ('降落高度(越小越高)',str(self.parameters.fall_height), '落地高度(越大越高)',str(self.parameters.land_height)),
            ('早间起始(h)', str(self.parameters.morning_start), '早间终止', str(self.parameters.morning_end)),
            ('午间起始', str(self.parameters.noon_start), '午间终止', str(self.parameters.noon_end))
        ]
        dialog=QDialog(self)
        dialog.setWindowTitle('调节参数(不填保持当前灰色值)')
        layout = QVBoxLayout()
        grid_layout = QGridLayout()
        dialog.input_boxes = []
        for row, (label1_text, label1_now, label2_text, label2_now) in enumerate(labels):
            # 创建标签
            label1 = QLabel(label1_text)
            label2 = QLabel(label2_text)
            # 创建输入文本框
            input_box1 = QLineEdit()
            input_box1.setPlaceholderText(label1_now)
            input_box2 = QLineEdit()
            input_box2.setPlaceholderText(label2_now)
            # 将标签和输入框添加到网格布局中
            grid_layout.addWidget(label1, row, 0)
            grid_layout.addWidget(input_box1, row, 1)
            grid_layout.addWidget(label2, row, 2)
            grid_layout.addWidget(input_box2, row, 3)
            dialog.input_boxes.extend([input_box1, input_box2])
        # 创建确定按钮
        ok_button = QPushButton('确定')
        ok_button.clicked.connect(dialog.accept)
        # 将网格布局和确定按钮添加到垂直布局中
        layout.addLayout(grid_layout)
        layout.addWidget(ok_button)
        dialog.setLayout(layout)
        if dialog.exec_()==QDialog.Accepted:
            inputs = [box.text() for box in dialog.input_boxes]
            self.set_volume(inputs[0])
            self.set_interval(inputs[1])
            self.set_fall_height(inputs[2])
            self.set_land_height(inputs[3])
            self.set_morning_start(inputs[4])
            self.set_morning_end(inputs[5])
            self.set_noon_start(inputs[6])
            self.set_noon_end(inputs[7])

        del dialog

    def mute_text(self):
        return "静音:"+ ('开' if self.status.is_muted else '关')

    def mute_setting(self):
        self.status.is_muted=not self.status.is_muted

    def random_voice_text(self):
        return "随机语音:" + ('开' if self.status.is_random_voice else '关')

    def random_voice_setting(self):
        self.status.is_random_voice= not self.status.is_random_voice
        if self.status.is_random_voice:
            self.timer2.start(self.parameters.timer2_interval)
        else:
            self.timer2.stop()

    def quit(self):
        global app
        config=read_config(self.config_file)
        config['MUTE'] = self.status.is_muted
        config['RANDOM_VOICE'] = self.status.is_random_voice
        config['VOLUME'] = self.parameters.volume
        config['FALL_HEIGHT'] = self.parameters.fall_height
        config['LAND_HEIGHT'] = self.parameters.land_height
        config['INTERVAL'] = self.parameters.timer2_interval
        config['MORNING_START'] = self.parameters.morning_start
        config['MORNING_END'] = self.parameters.morning_end
        config['NOON_START'] = self.parameters.noon_start
        config['NOON_END'] = self.parameters.noon_end
        save_config(config, self.config_file)
        app.quit()

    def day_count_text(self):
        return f"已陪伴您: {self.memory_day.day_count} 天"

    def menu_add(self,menu,text,func=None):
        action=QAction(text,self)
        if func is not None:
            action.triggered.connect(func)
        menu.addAction(action)

    def show_menu(self):
        menu=QMenu(self)

        self.menu_add(menu,self.day_count_text())
        self.menu_add(menu,self.mute_text(),self.mute_setting)
        self.menu_add(menu,self.random_voice_text(),self.random_voice_setting)
        self.menu_add(menu,self.parameters_text(),self.parameters_setting)
        self.menu_add(menu,'退出',self.quit)

        menu.exec(QCursor.pos())

        del menu

    def press_to_menu(self):
        self.status.is_walking = False
        self.status.is_falling = False
        self.status.is_dragging = False
        self.status.is_clicking = False
        self.show_menu()

def on_quit(icon,item):
    global app
    global pet
    config=read_config(pet.config_file)
    config['MUTE']=pet.status.is_muted
    config['RANDOM_VOICE']=pet.status.is_random_voice
    config['VOLUME']=pet.parameters.volume
    config['FALL_HEIGHT']=pet.parameters.fall_height
    config['LAND_HEIGHT']=pet.parameters.land_height
    config['INTERVAL']=pet.parameters.timer2_interval
    config['MORNING_START']=pet.parameters.morning_start
    config['MORNING_END']=pet.parameters.morning_end
    config['NOON_START']=pet.parameters.noon_start
    config['NOON_END']=pet.parameters.noon_end
    save_config(config,pet.config_file)
    icon.stop()
    app.quit()

menu=(
    pystray.MenuItem('退出', on_quit),
)

app = QApplication(sys.argv)
pet = DesktopPet()
pet.show()

icon_path=os.path.join(pet.current_dir,'source','应用图标','icon.png')
image=Image.open(icon_path)
icon=pystray.Icon('Flutterpage',image,'Flutterpage',menu)

icon_thread=threading.Thread(target=icon.run)
icon_thread.daemon=True
icon_thread.start()

sys.exit(app.exec_())

