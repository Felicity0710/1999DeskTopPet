import os.path
import sys
import random
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QSizePolicy
from PyQt5.QtGui import QMovie
from PyQt5.QtCore import Qt, QTimer,QUrl
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
import pystray
from PIL import Image
from pathlib import Path
import tkinter as tk
from tkinter import simpledialog
import datetime
import json
from openpyxl import load_workbook

nan_strings = ["nan", "NaN", "NAN"]
MMMute=False
current_dir=os.path.dirname(os.path.abspath(__file__))
gif_dir=os.path.join(current_dir,'source','动作gif格式')
voice_dir=os.path.join(current_dir,'source','播放语音素材')
config_file = os.path.join(current_dir, 'config.json')
born_gif_dir=Path(os.path.join(gif_dir,'出生动作'))
drag_gif_dir=Path(os.path.join(gif_dir,'拖拽动作'))
idle_gif_dir=Path(os.path.join(gif_dir,'待机动作'))
click_gif_dir=Path(os.path.join(gif_dir,'点击动作'))
walk_gif_dir=Path(os.path.join(gif_dir,'行走动作'))
fly_gif_dir=Path(os.path.join(gif_dir,'飞行动作'))
bg=[file.name for file in born_gif_dir.iterdir() if file.is_file()]
dg=[file.name for file in drag_gif_dir.iterdir() if file.is_file()]
ig=[file.name for file in idle_gif_dir.iterdir() if file.is_file()]
cg=[file.name for file in click_gif_dir.iterdir() if file.is_file()]
wg=[file.name for file in walk_gif_dir.iterdir() if file.is_file()]
fg=[file.name for file in fly_gif_dir.iterdir() if file.is_file()]
workbook=load_workbook(os.path.join(voice_dir,'语音目录.xlsx'))
worksheet=workbook['Sheet1']
def read_config():
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
                return config
        except json.JSONDecodeError:
            pass
    return {}
def save_config(config):
    with open(config_file, 'w') as f:
        json.dump(config, f)

def getday():
    now=datetime.datetime.now()
    hour=now.hour
    if 5 <= hour < 12:
        return 0
    elif 12<=hour<18:
        return 1
    else:
        return 2

class DesktopPet(QWidget):

    #初始化
    def __init__(self):
        super().__init__()
        #一些状态参数
        self.is_dragging = False
        self.drag_offset = None
        self.falling = False
        self.walking = False
        self.walk_direction = -1
        self.walk_speed = 2
        self.fall_speed = 5
        self.prev_state = None
        self.prev_mouse_pos = None  # 新增：记录上一次鼠标位置
        self.is_muted=False
        self.volume=50
        self.random_position()
        self.reserve_space=50
        self.reserve_space2=100
        self.timer2_interval=12000
        self.timer_interval=25
        self.is_random_voice=True
        self.fall_height=250
        self.land_height=5

        #基础设置
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.current_animation = None

        #动画刷新Timer
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_animation)
        self.timer.start(self.timer_interval)

        #音频播放器
        self.player = QMediaPlayer(self)
        self.player.setVolume(self.volume)

        #动画素材
        self.drag_gif = [os.path.join(gif_dir,'拖拽动作',it) for it in dg]
        self.fall_gif = [os.path.join(gif_dir,'飞行动作',it) for it in fg]
        self.walk_gif = [os.path.join(gif_dir,'行走动作',it) for it in wg]
        self.special_gif = [os.path.join(gif_dir,'点击动作',it) for it in cg]
        self.born_gif=[os.path.join(gif_dir,'出生动作',it)for it in bg]
        self.idle_gifs = [os.path.join(gif_dir, '待机动作', it) for it in ig ]
        self.drag_gg=random.choice(self.drag_gif)
        self.fall_gg=random.choice(self.fall_gif)
        self.walk_gg=random.choice(self.walk_gif)
        #台词，语音，语音长度素材
        self.idle_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[0] is None:
                    break
                else:
                    if row[1] is None:
                        self.idle_word.append((row[0].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.idle_word.append((row[0],os.path.join(voice_dir,row[1]),2500))
        self.pre_morning_idle_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[2] is None:
                    break
                else:
                    if row[3] is None:
                        self.pre_morning_idle_word.append((row[2].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.pre_morning_idle_word.append((row[2].replace('\\n','\n').strip(),os.path.join(voice_dir,row[3].strip()),2500))
        self.pre_noon_idle_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[4] is None:
                    break
                else:
                    if row[5] is None:
                        self.pre_noon_idle_word.append((row[4].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.pre_noon_idle_word.append((row[4].replace('\\n','\n').strip(),os.path.join(voice_dir,row[5].strip()),2500))
        self.pre_night_idle_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[6] is None:
                    break
                else:
                    if row[7] is None:
                        self.pre_night_idle_word.append((row[6].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.pre_night_idle_word.append((row[6].replace('\\n','\n').strip(),os.path.join(voice_dir,row[7].strip()),2500))
        self.click_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[8] is None:
                    break
                else:
                    if row[9] is None:
                        self.click_word.append((row[8].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.click_word.append((row[8].replace('\\n','\n').strip(),os.path.join(voice_dir,row[9].strip()),2500))
        self.drag_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[10] is None:
                    break
                else:
                    if row[11] is None:
                        self.drag_word.append((row[10].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.drag_word.append((row[10].replace('\\n','\n').strip(),os.path.join(voice_dir,row[11].strip()),2500))
        self.fall_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[12] is None:
                    break
                else:
                    if row[13] is None:
                        self.fall_word.append((row[12].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.fall_word.append((row[12].replace('\\n','\n').strip(),os.path.join(voice_dir,row[13].strip()),2500))
        self.born_word=[]
        for row in worksheet.iter_rows(min_row=2,values_only=True):
            if row:
                if row[14] is None:
                    break
                else:
                    if row[15] is None:
                        self.born_word.append((row[14].replace('\\n', '\n').strip(), voice_dir, 2500))
                    else:
                        self.born_word.append((row[14].replace('\\n','\n').strip(),os.path.join(voice_dir,row[15].strip()),2500))
        self.morning_idle_word=self.idle_word+self.pre_morning_idle_word
        self.noon_idle_word = self.idle_word+self.pre_noon_idle_word
        self.night_idle_word= self.idle_word+self.pre_night_idle_word
        self.dword=random.choice(self.drag_word)
        self.fword=random.choice(self.fall_word)
        #对话框属性
        self.dialog_label = QLabel(self)
        self.dialog_label.setStyleSheet("""
            background-color: rgba(0, 0, 0, 0);
            color: white;
            font-family: "Segoe UI", sans-serif;
            font-size: 12pt;
            font-weight: bold;
        """)
        self.dialog_label.setAlignment(Qt.AlignCenter)
        self.dialog_label.setVisible(False)  # 默认不显示

        #随机语音Timer
        self.timer2 = QTimer(self)
        self.timer2.timeout.connect(self.display_conversation)
        self.timer2.start(self.timer2_interval)

        #出生动画设置
        self.current_movie = self.load_gif(random.choice(self.born_gif)) #出生动画
        self.current_movie.frameChanged.connect(self.check_last_frame)
        self.label = QLabel(self)
        self.label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)
        self.label.setMovie(self.current_movie)

        #出生随机语音设置
        self.timer2.stop()
        bword=random.choice(self.born_word)
        self.dialog_label.setText(bword[0])
        animation_rect = self.label.geometry()
        dialog_x = animation_rect.center().x()
        dialog_y = self.dialog_label.y()
        self.dialog_label.move(dialog_x, dialog_y)
        self.dialog_label.setVisible(True)
        if not MMMute:
            path = bword[1]
            if path != voice_dir:
                audio_file = QUrl.fromLocalFile(path)
                self.player.setMedia(QMediaContent(audio_file))
                self.player.play()
        # 设置台词若干秒后消失
        QTimer.singleShot(bword[2], self.hide_conversation)
        self.restart_timer2()

    #加载动画图片
    def load_gif(self, gif_path):
        movie = QMovie(gif_path)
        movie.setCacheMode(QMovie.CacheAll)
        movie.start()
        frame = movie.currentPixmap()
        width = frame.width()
        height = frame.height()
        self.setFixedSize(width+self.reserve_space2, height+self.reserve_space)
        return movie

    #随机出生坐标
    def random_position(self):
        screen_geometry = QApplication.desktop().screenGeometry()
        x = random.randint(0, screen_geometry.width() - self.width())
        y = random.randint(0, screen_geometry.height() - self.height())
        self.move(x, y)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.is_dragging = True
            self.drag_offset = event.pos()
            self.prev_mouse_pos = event.globalPos()  # 记录按下时的鼠标全局位置
            self.timer2.stop()
            self.dword=random.choice(self.drag_word)
            self.drag_gg=random.choice(self.drag_gif)
            if not self.is_muted:
                path = self.dword[1]
                if path!=voice_dir:
                    audio_file = QUrl.fromLocalFile(path)
                    self.player.setMedia(QMediaContent(audio_file))
                    self.player.play()
        elif event.button()==Qt.RightButton:
            self.walking=False
            self.falling=False

    #拖拽专属设置
    def mouseMoveEvent(self, event):
        if self.is_dragging:
            current_mouse_pos = event.globalPos()
            if self.prev_mouse_pos:
                # 计算鼠标移动的偏移量
                delta = current_mouse_pos - self.prev_mouse_pos
                new_pos = self.pos() + delta

                screens = QApplication.screens()
                combined_geometry = screens[0].geometry()
                for screen in screens[1:]:
                    combined_geometry = combined_geometry.united(screen.geometry())

                new_x = max(combined_geometry.left(), min(new_pos.x(), combined_geometry.right() - self.width()))
                new_y = max(combined_geometry.top(), min(new_pos.y(), combined_geometry.bottom() - self.height()))

                self.move(new_x, new_y)
                self.prev_mouse_pos = current_mouse_pos  # 更新上一次鼠标位置

                # 拖拽动作
                self.prev_state = self.get_current_state()
                self.current_movie.stop()
                self.label.clear()
                new_movie = self.load_gif(self.drag_gg)  # 这里设置拖拽动作
                self.current_movie = new_movie
                self.current_movie.frameChanged.connect(self.restore_previous_state)
                self.label.setMovie(self.current_movie)
                self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
                self.current_movie.start()
                self.update()

                #拖拽台词
                self.dialog_label.setText(self.dword[0])
                animation_rect = self.label.geometry()
                dialog_x = animation_rect.center().x() - self.dialog_label.width() // 2 - 10
                dialog_y = self.dialog_label.y()
                self.dialog_label.move(dialog_x, dialog_y)
                self.dialog_label.setVisible(True)

    #拖拽结束/点击结束设置
    def mouseReleaseEvent(self, event):
        if event.button() == Qt.RightButton:
            #点击台词和语音
            self.timer2.stop()
            cword=random.choice(self.click_word)
            self.dialog_label.setText(cword[0])
            animation_rect = self.label.geometry()
            dialog_x = animation_rect.center().x() - self.dialog_label.width() // 2 - 10
            dialog_y = self.dialog_label.y()
            self.dialog_label.move(dialog_x, dialog_y)
            self.dialog_label.setVisible(True)
            if not self.is_muted:
                path =cword[1]
                if path != voice_dir:
                    audio_file = QUrl.fromLocalFile(path)
                    self.player.setMedia(QMediaContent(audio_file))
                    self.player.play()
            QTimer.singleShot(cword[2], self.hide_conversation)
            QTimer.singleShot(cword[2],self.restart_timer2)
            #点击动作
            self.prev_state = self.get_current_state()
            self.current_movie.stop()
            self.label.clear()
            new_movie = self.load_gif(random.choice(self.special_gif)) #这里设置点击动作
            self.current_movie = new_movie
            self.current_movie.frameChanged.connect(self.restore_previous_state)
            self.label.setMovie(self.current_movie)
            self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
            self.update()
        elif event.button() == Qt.LeftButton:
            self.is_dragging = False
            self.prev_mouse_pos = None
            self.dialog_label.setVisible(False)
            self.restart_timer2()
            #下落高度及其动作
            self.fall_gg = random.choice(self.fall_gif)
            if self.pos().y() < self.fall_height:
                self.falling = True
                if not self.is_muted:
                    path = self.fword[1]
                    if path != voice_dir:
                        audio_file = QUrl.fromLocalFile(path)
                        self.player.setMedia(QMediaContent(audio_file))
                        self.player.play()
                self.current_movie.stop()
                self.label.clear()
                new_movie = self.load_gif(self.fall_gg) #这里设置下落动作
                self.current_movie = new_movie
                self.current_movie.frameChanged.connect(self.check_last_frame)
                self.label.setMovie(self.current_movie)
                self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
                self.update()
            #不下落就恢复idle状态
            else:
                self.current_movie.stop()
                self.label.clear()
                current_pos = self.pos()
                screen_geometry = QApplication.desktop().screenGeometry()
                left_bottom_y = current_pos.y() + self.height()
                new_movie = self.load_gif(random.choice(self.idle_gifs))
                self.current_movie = new_movie
                self.current_movie.frameChanged.connect(self.check_last_frame)
                self.label.setMovie(self.current_movie)
                self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
                self.update()
                new_y = left_bottom_y - self.height()
                self.move(current_pos.x(), new_y)

    #下落/行走设置
    def update_animation(self):
        if self.falling and not self.is_dragging:
            if self.timer2.isActive():
                self.timer2.stop()
            new_y = self.pos().y() + self.fall_speed
            screen_geometry = QApplication.desktop().screenGeometry()
            #下落台词/语音设置
            self.fword=random.choice(self.fall_word)
            self.dialog_label.setText(self.fword[0])
            animation_rect = self.label.geometry()
            dialog_x = animation_rect.center().x() - self.dialog_label.width() // 2 - 10
            dialog_y = self.dialog_label.y()
            self.dialog_label.move(dialog_x, dialog_y)
            self.dialog_label.setVisible(True)
            #落到底端开始行走
            self.walk_gg=random.choice(self.walk_gif)
            if new_y >= screen_geometry.height() - self.height() -self.land_height:
                new_y = screen_geometry.height() - self.height() -self.land_height
                self.falling = False
                #关闭下落台词
                self.dialog_label.setVisible(False)
                self.restart_timer2()
                self.walking = True
                self.current_movie.stop()
                self.label.clear()
                new_movie = self.load_gif(self.walk_gg)
                self.current_movie = new_movie
                self.current_movie.frameChanged.connect(self.check_last_frame)
                self.label.setMovie(self.current_movie)
                self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
                self.update()
            current_pos = self.pos()
            self.move(current_pos.x(), new_y)

        elif self.walking:
            new_x = self.pos().x() + self.walk_speed * self.walk_direction
            screen_geometry = QApplication.desktop().screenGeometry()
            self.move(new_x, self.pos().y())
            if new_x <= 20 or self.pos().y() < screen_geometry.height() - self.height() - self.land_height-45:
                self.walking = False
                self.current_movie.stop()
                self.label.clear()
                #行走结束后的动作设置
                current_pos = self.pos()
                left_bottom_y = current_pos.y() + self.height()
                new_movie = self.load_gif(random.choice(self.idle_gifs))
                self.current_movie = new_movie
                self.current_movie.frameChanged.connect(self.check_last_frame)
                self.label.setMovie(self.current_movie)
                self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
                new_y = left_bottom_y - self.height()
                self.move(current_pos.x(), new_y)
                self.update()

    #动作切换
    def check_last_frame(self, frame_number):
        if frame_number == self.current_movie.frameCount() - 1  and not self.falling and not self.walking:
            self.current_movie.stop()
            self.label.clear()
            current_pos = self.pos()
            left_bottom_y = current_pos.y() + self.height()
            new_movie = self.load_gif(random.choice(self.idle_gifs))
            self.current_movie = new_movie
            self.current_movie.frameChanged.connect(self.check_last_frame)
            self.label.setMovie(self.current_movie)
            self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
            new_y = left_bottom_y - self.height()
            self.move(current_pos.x(), new_y)
            self.update()

    #动作切换
    def restore_previous_state(self, frame_number):
        if frame_number == self.current_movie.frameCount() - 1:
            self.current_movie.stop()
            self.label.clear()
            current_pos = self.pos()
            left_bottom_y = current_pos.y() + self.height()  # 左下角y坐标
            if self.prev_state == 'idle':
                new_movie = self.load_gif(random.choice(self.idle_gifs))
            elif self.prev_state == 'drag':
                new_movie = self.load_gif(self.drag_gif)
            elif self.prev_state == 'fall':
                new_movie = self.load_gif(self.fall_gif)
            elif self.prev_state == 'walk':
                new_movie = self.load_gif(self.walk_gif)
            self.current_movie = new_movie
            self.current_movie.frameChanged.connect(self.check_last_frame)
            self.label.setMovie(self.current_movie)
            self.label.setGeometry(0, 0, new_movie.currentPixmap().width()+self.reserve_space2, new_movie.currentPixmap().height()+self.reserve_space)
            new_y = left_bottom_y - self.height()
            self.move(current_pos.x(), new_y)
            self.update()
            self.prev_state = None

    #获取状态
    def get_current_state(self):
        if self.is_dragging:
            return 'drag'
        elif self.falling:
            return 'fall'
        elif self.walking:
            return 'walk'
        else:
            return 'idle'

    #idle状态随机台词/语音
    def display_conversation(self):
        if not self.is_dragging and not self.falling:
            day=getday()
            if day==0:
                word_voice=random.choice(self.morning_idle_word)
            elif day==1:
                word_voice=random.choice(self.noon_idle_word)
            else:
                word_voice=random.choice(self.night_idle_word)
            conversation_text = word_voice[0]
            self.dialog_label.setText(conversation_text)
            animation_rect = self.label.geometry()
            dialog_x = animation_rect.center().x() - self.dialog_label.width() // 2-10
            dialog_y = self.dialog_label.y()
            self.dialog_label.move(dialog_x, dialog_y)
            self.dialog_label.setVisible(True)
            #如果没静音
            if not self.is_muted:
                path=word_voice[1]
                if path != voice_dir:
                    audio_file = QUrl.fromLocalFile(path)
                    self.player.setMedia(QMediaContent(audio_file))
                    self.player.play()
            #设置台词若干秒后消失
            QTimer.singleShot(word_voice[2], self.hide_conversation)
    #隐藏台词
    def hide_conversation(self):
        self.dialog_label.setVisible(False)  # 隐藏对话框

    def restart_timer2(self):
        if self.is_random_voice:
            self.timer2.start(self.timer2_interval)

config = read_config()
if 'Mute' in config:
    MMMute=config['Mute']
app = QApplication(sys.argv)
pet = DesktopPet()
if 'Mute' in config:
    pet.is_muted = config['Mute']
if 'Volume' in config:
    pet.volume=config['Volume']
    pet.player.setVolume(pet.volume)
if 'Interval' in config:
    pet.timer2_interval=config['Interval']
    pet.timer2.start(pet.timer2_interval)
if 'Fall_height' in config:
    pet.fall_height=config['Fall_height']
if 'Land_height' in config:
    pet.land_height=config['Land_height']
if 'Random_voice' in config:
    pet.is_random_voice=config['Random_voice']
    if not pet.is_random_voice:
        pet.timer2.stop()

pet.show()
def on_quit(icon,item):
    global app
    icon.stop()
    app.quit()
def mute(icon,item):
    global pet
    config = read_config()
    if pet.is_muted:
        pet.is_muted=False
        config['Mute']=False
    else:
        pet.is_muted=True
        config['Mute']=True
    save_config(config)
def random_voice(icon,item):
    global pet
    config=read_config()
    if pet.timer2.isActive():
        pet.timer2.stop()
        pet.is_random_voice=False
        config['Random_voice']=False
    else:
        pet.is_random_voice=True
        config['Random_voice']=True
        pet.timer2.start(pet.timer2_interval)
    save_config(config)
def set_volume_by_input(icon, item):
    def tk_input_dialog():
        root = tk.Tk()
        root.withdraw()
        try:
            volume_value = simpledialog.askfloat("设置音量", "请输入音量值 (0 - 100):", minvalue=0, maxvalue=100)
            if volume_value is not None:
                pet.volume = int(volume_value)
                pet.player.setVolume(pet.volume)
                config = read_config()
                config['Volume']=pet.volume
                save_config(config)
        except ValueError:
            print("输入的不是有效的数值，请输入 0 到 100 之间的数字。")
        finally:
            root.destroy()
        icon.update_menu()
    tk_thread = threading.Thread(target=tk_input_dialog)
    tk_thread.start()
def set_intervals_by_input(icon,item):
    def tk_input_dialog():
        root = tk.Tk()
        root.withdraw()
        try:
            interval_value = simpledialog.askfloat("设置时间间隔", "请输入时间间隔(s):", minvalue=0,)
            if interval_value is not None:
                pet.timer2_interval=int(interval_value*1000)
                config = read_config()
                config['Interval']=pet.timer2_interval
                save_config(config)
                if pet.timer2.isActive():
                    pet.timer2.stop()
                    pet.restart_timer2()
        finally:
            root.destroy()
        icon.update_menu()
    tk_thread = threading.Thread(target=tk_input_dialog)
    tk_thread.start()
def set_fall_height_by_input(icon,item):
    def tk_input_dialog():
        root = tk.Tk()
        root.withdraw()
        try:
            height_value = simpledialog.askfloat("设置降落高度", "请输入高度（越小越高）:", minvalue=0,)
            if height_value is not None:
                pet.fall_height=int(height_value)
                config = read_config()
                config['Fall_height'] = pet.fall_height
                save_config(config)
        finally:
            root.destroy()
        icon.update_menu()
    tk_thread = threading.Thread(target=tk_input_dialog)
    tk_thread.start()
def set_land_height_by_input(icon,item):
    def tk_input_dialog():
        root = tk.Tk()
        root.withdraw()
        try:
            height_value = simpledialog.askfloat("设置落地高度", "请输入高度（越大越高）:")
            if height_value is not None:
                pet.land_height=int(height_value)
                config = read_config()
                config['Land_height'] = pet.land_height
                save_config(config)
        finally:
            root.destroy()
        icon.update_menu()
    tk_thread = threading.Thread(target=tk_input_dialog)
    tk_thread.start()
def volume_text(item):
    return f"音量调节(当前:{pet.volume})"
def voice_text(item):
    return f'语音时间间隔(当前:{int(pet.timer2_interval/1000)}s)'
def fall_text(item):
    return f"降落起始高度(当前:{int(pet.fall_height)})"
def land_text(item):
    return f"落地终止高度(当前:{int(pet.land_height)})"
icon_path=os.path.join(current_dir,'source','应用图标','icon.png')
image=Image.open(icon_path)
#菜单
menu=(
    #静音菜单
    pystray.MenuItem(
        lambda item: f"静音:{'开' if pet.is_muted else '关'}",
    mute),
    #随机语音菜单
    pystray.MenuItem(lambda item: f"随机语音:{'开' if pet.is_random_voice else '关'}",
                     random_voice),
    #音量菜单
    pystray.MenuItem(volume_text,
                     set_volume_by_input
    ),
    #随机语音间隔时间菜单
    pystray.MenuItem(voice_text,
                     set_intervals_by_input
    ),
    #降落起始高度
    pystray.MenuItem(fall_text,
                     set_fall_height_by_input
    ),
    #落地终止高度
    pystray.MenuItem(land_text,
                     set_land_height_by_input
    ),

    pystray.MenuItem('退出',on_quit),
)
icon=pystray.Icon('Flutterpage',image,'Flutterpage',menu)

import threading
icon_thread=threading.Thread(target=icon.run)
icon_thread.daemon=True
icon_thread.start()

sys.exit(app.exec_())
